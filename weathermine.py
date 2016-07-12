import json
import requests
import csv
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

#API Key for future reference
api_key = "3715bd6fcab875b08dc6691db0412e20"

#make the call to the API
url = "http://api.openweathermap.org/data/2.5/forecast/city?id=524901&APPID=3715bd6fcab875b08dc6691db0412e20"
response = requests.get(url)
#Invoke if there are any errors
response.raise_for_status()

#Dump the data into a JSON file in case need to review or acess it
with open('weather.json', 'w') as outfile:
     json.dump(response.json(), outfile, indent=4)


# Load JSON data into a Python variable.
weatherData = json.loads(response.text)


#Make into local list for easy manipulation
w = weatherData['list']


#open csv file that will be wrtitten too 
outputFile = open('output.csv', 'w', newline='')
outputWriter = csv.writer(outputFile)

#Line Headers
outputWriter.writerow(['Date', 'Temp', 'Min Temp', 'Max Temp'])

#Collect 7 days worth of data from the JSON file and write it into the CSV file 
for i in range (7):
    h = [datetime.date.today() + datetime.timedelta(days=i), w[i]['main']['temp'], w[i]['main']['temp_min'], w[i]['main']['temp_max']]
    outputWriter.writerow(h)

#Close the file 
outputFile.close()


#Reopen the file and create a CSV reader object 
examplefile = open('output.csv')
exampleReader = csv.reader(examplefile)
exampleData = list(exampleReader)


#Create an excel workbook to transfer the data into
wb = openpyxl.Workbook()
ws = wb.get_sheet_by_name('Sheet')

#Transfer title headers
for i in range (1, 5):
    a = i-1
    ws.cell(row=1, column=i).value = exampleData[0][a]

#Transfer the dates
for i in range (1, 9):
    a = i-1
    ws.cell(row=i, column=1).value = exampleData[a][0]

#Transfer the temps (needs to be separate in order to parse these into floats)
for i in range (2,9):
    for h in range(2, 5):
        a = i-1
        b = h-1
        ws.cell(row=i, column=h).value = float(exampleData[a][b])

ws = wb.active

#Create Chart object
c1 = LineChart()

#Edit Chart details 
c1.title = "Weekly Temps"
c1.y_axis_title = "Temp"
c1.x_axis_title = "Date"
c1.height = 20
c1.width = 40

#Collect Reference points for the chart
data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=8)
dates = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=8)

#Add Everything to the chart
c1.add_data(data, titles_from_data=True)
c1.set_categories(dates)

#Add Chart to sheet
ws.add_chart(c1)

#Save
wb.save('WeeklyWeather.xlsx')

